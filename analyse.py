import json
import openpyxl
from openpyxl.styles.borders import Border, Side
import os.path as path
import os
import pandas as pd
import pprint
from tkinter import filedialog as fd
import ciaReport

fileDir = ""
fileName = ""
filePath = ""
sheetName = ""
analysisFilePath = ""
xlFile = ""
outputDir = ""


def setFilePath():
    global fileDir, fileName, filePath
    filetypes = [
        ('Excel Files', '*.xlsx')
    ]
    filePath = fd.askopenfilename(title='Open internal marks excel file', initialdir=str(os.getenv("HOME")),
                                  filetypes=filetypes).replace("\\", "/")
    fileName = path.basename(filePath)
    fileDir = path.dirname(filePath)
    return 1


def getPercentage(marks):
    if marks == "NaN":
        return "N/A"
    elif marks != -1:
        return round((marks / 40) * 100, 2)
    else:
        return "N/A"


def getGrade(percent):
    if percent != "N/A":
        if 70 <= percent <= 100:
            return "FCD"
        # 60 <= marks <= 69, Grade="FC", GPA=Total/9.5 rounded to two decimal places.
        elif 60 <= percent <= 69:
            return "FC"
        # 50 <= marks <= 59, Grade="SC", GPA=Total/9.5 rounded to two decimal places.
        elif 50 <= percent <= 59:
            return "SC"
        else:
            return "F"
    else:
        return "N/A"


# noinspection PyBroadException
def getSheetName():
    global filePath, fileName, fileDir, sheetName, outputDir
    wb = openpyxl.open(filePath, read_only=True, keep_links=False)
    sheetName = wb.sheetnames[0]
    try:
        os.makedirs(os.path.join(fileDir, sheetName), exist_ok=True)
    except:
        print("Couldn't create directory !!!\n", outputDir)
        print("Exiting...")
        exit(4)
    outputDir = os.path.join(fileDir, sheetName)


# noinspection PyBroadException,PyUnresolvedReferences
def toJson():
    global fileDir, fileName, filePath, sheetName, outputDir
    # only usn, name and subject columns are selected and worked with
    # change "usecols" value to the desired columns if number of subject changes
    # format--> "firstCol:lastCol"
    try:
        internalFrame = pd.read_excel(filePath, sheet_name=sheetName, header=0, usecols="B:K", skiprows=2)
    except:
        internalFrame = pd.read_excel(filePath, sheet_name=sheetName, header=0, usecols="B:K", skiprows=2, engine="openpyxl")

    studentDict = {}
    # print(internalFrame.columns)
    colList = list(internalFrame.keys())
    colList.remove("USN")
    colList.remove("NAME")
    try:
        colList.remove("Section")
    except:
        print("Section column not added!\n Add \"Section\" columns and run the program again")
    marksDict = {}
    for i in range(0, len(internalFrame["USN"])):
        studentDict[internalFrame["USN"][i]] = {}
        for col in colList:
            try:
                newCol = col.split("(")[0]
            except:
                newCol = col
            if not str(internalFrame[col][i]) == "nan":
                marksDict[str(newCol).strip(" ")] = {}
                if str(internalFrame[col][i]) in "ABabAbaBaaAA" or str(internalFrame[col][i]) == "0":
                    marksDict[str(newCol).strip(" ")]["marks"] = -1
                else:
                    marksDict[str(newCol).strip(" ")]["marks"] = int(internalFrame[col][i])
                marksDict[str(newCol).strip(" ")]["percentage"] = getPercentage(marksDict[str(newCol).strip(" ")]["marks"])
                marksDict[str(newCol).strip(" ")]["grade"] = getGrade(marksDict[str(newCol).strip(" ")]["percentage"])
        studentDict[internalFrame["USN"][i]]["Name"] = internalFrame["NAME"][i]
        studentDict[internalFrame["USN"][i]]["Section"] = internalFrame["Section"][i]
        studentDict[internalFrame["USN"][i]]["marks"] = marksDict
        marksDict = {}
    # pprint.pprint(studentDict, indent=2, sort_dicts=False)

    for i in studentDict.keys():
        analysis = {"percentage": 0.0}
        for sub in studentDict[i]["marks"].keys():
            try:
                analysis["percentage"] += studentDict[i]["marks"][sub]["percentage"]
            except TypeError:
                continue
        analysis["percentage"] = round(analysis["percentage"] / (len(studentDict[i]["marks"].keys()) - 1), 2)
        analysis["grade"] = getGrade(analysis["percentage"])
        studentDict[i]["analysis"] = analysis
    # pprint.pprint(studentDict, indent=2, sort_dicts=False)
    with open(path.join(outputDir, "studentMarks.json"), "w") as file:
        file.write(json.dumps(studentDict, indent=4))

    subList = []
    for usn in studentDict.keys():
        for sub in studentDict[usn]["marks"].keys():
            if sub not in subList:
                subList.append(sub)
    # subList.remove("Name")

    subAnalysis = {}
    subMarks = {}
    for sub in subList:
        if sub not in subAnalysis.keys():
            subAnalysis[sub] = {"FCD": 0, "FC": 0, "SC": 0, "F": 0, "AB": 0, "TOTAL_APPEARED": 0, "MARKS": {},
                                "TOP_1": "", "TOP_2": "", "TOP_3": ""}
            subMarks[sub] = {}

    for usn in studentDict.keys():
        for sub in subList:
            try:
                if sub in studentDict[usn]["marks"].keys():
                    subAnalysis[sub]["TOTAL_APPEARED"] += 1
                    subAnalysis[sub]["TOP_1"] = usn
                    subAnalysis[sub]["TOP_2"] = usn
                    subAnalysis[sub]["TOP_3"] = usn
                    subAnalysis[sub]["MARKS"][usn] = studentDict[usn]["marks"][sub]["marks"]
                    if studentDict[usn]["marks"][sub]["grade"] == "FCD":
                        subAnalysis[sub]["FCD"] += 1
                    if studentDict[usn]["marks"][sub]["grade"] == "FC":
                        subAnalysis[sub]["FC"] += 1
                    if studentDict[usn]["marks"][sub]["grade"] == "SC":
                        subAnalysis[sub]["SC"] += 1
                    if studentDict[usn]["marks"][sub]["grade"] == "F":
                        subAnalysis[sub]["F"] += 1
                    if studentDict[usn]["marks"][sub]["marks"] == -1:
                        subAnalysis[sub]["AB"] += 1

                    subMarks[sub][usn] = studentDict[usn]["marks"][sub]["marks"]
            except:
                continue

    for sub in subMarks.keys():
        sortedUSN = sorted(subMarks[sub], key=subMarks[sub].__getitem__, reverse=True)
        for i in range(1, 4):
            top_string = "TOP_" + str(i)
            subAnalysis[sub][top_string] = sortedUSN[i]

    # pprint.pprint(subAnalysis, indent=2, sort_dicts=False)
    with open(path.join(outputDir, "subjectAnalysis.json"), "w") as file:
        file.write(json.dumps(subAnalysis, indent=4))
    return 1


# noinspection PyShadowingNames,PyBroadException
def markToXl():
    global analysisFilePath
    with open(path.join(outputDir, "studentMarks.json"), "r") as file:
        studentJson = json.load(file)
    # pprint.pprint(studentJson, indent=2, sort_dicts=False)
    mainDict = {"USN": [], "Name": []}
    subList = []
    for student in studentJson.keys():
        for sub in studentJson[student]["marks"].keys():
            if sub not in subList:
                subList.append(sub)
    for sub in subList:
        mainDict[str(sub + "_marks")] = []
        mainDict[str(sub + "_percentage")] = []
        mainDict[str(sub + "_grade")] = []

    for student in studentJson.keys():
        mainDict["USN"].append(student)
        mainDict["Name"].append(studentJson[student]["Name"])
        for sub in subList:
            try:
                marks = int(studentJson[student]["marks"][sub]["marks"])
                if marks == -1:
                    mainDict[str(sub + "_marks")].append("AB")
                else:
                    mainDict[str(sub + "_marks")].append(marks)
            except KeyError:
                mainDict[str(sub + "_marks")].append("")
            try:
                mainDict[str(sub + "_percentage")].append(studentJson[student]["marks"][sub]["percentage"])
            except KeyError:
                mainDict[str(sub + "_percentage")].append("")
            try:
                mainDict[str(sub + "_grade")].append(studentJson[student]["marks"][sub]["grade"])
            except KeyError:
                mainDict[str(sub + "_grade")].append("")
    resultFrame = pd.DataFrame(mainDict)
    # with pd.option_context('display.max_rows', None, 'display.max_columns', None):
    #     print(resultFrame)
    analysisFileName = str(sheetName) + "_Analysis.xlsx"
    analysisFilePath = path.join(outputDir, analysisFileName).replace("\\", "/")
    xlFile = analysisFilePath
    try:
        xlFileWriter = pd.ExcelWriter(xlFile, engine="openpyxl")
        print("\tUsing \"openpyxl\" engine to read file")
    except:
        xlFileWriter = pd.ExcelWriter(xlFile)
    resultFrame.to_excel(xlFileWriter, sheet_name="StudentAnalysis", header=True, index=False, index_label=None)

    with open(path.join(outputDir, "subjectAnalysis.json"), "r") as file:
        subjectJson = json.load(file)
    for sub in subjectJson.keys():
        newDict = {}
        # tempDict = dict(subjectJson[sub])
        # for i in tempDict.keys():
        #     value = tempDict[i]
        #     newDict[i] = [value]
        tempDataFrame = pd.DataFrame(newDict)
        tempDataFrame.to_excel(xlFileWriter, sheet_name=sub, header=True, index=False, index_label=None)
    xlFileWriter.save()
    return 1


# noinspection PyUnboundLocalVariable,PyBroadException
def listStudents():
    global fileDir, sheetName, analysisFilePath
    try:
        workBook = openpyxl.load_workbook(analysisFilePath)
    except FileNotFoundError:
        print("file not found in ", fileDir)
        exit(2)
    # noinspection PyUnboundLocalVariable
    sheetNames = workBook.sheetnames
    sheetNames.remove('StudentAnalysis')
    subjectWiseSelection = {}
    for sub in sheetNames:
        subjectWiseSelection[sub] = {"FCD": {}, "FC": {}, "SC": {}, "F": {}, "AB": {}}
    with open(path.join(outputDir, "studentMarks.json"), "r") as file:
        studentJson = json.load(file)
    with open(path.join(outputDir, "subjectAnalysis.json"), "r") as file:
        subjectJson = json.load(file)

    for usn in studentJson.keys():
        for sub in sheetNames:
            try:
                grade = studentJson[usn]["marks"][sub]["grade"]
            except KeyError:
                continue
            if grade == "FCD":
                subjectWiseSelection[sub]["FCD"][usn] = studentJson[usn]["Name"]
            elif grade == "FC":
                subjectWiseSelection[sub]["FC"][usn] = studentJson[usn]["Name"]
            elif grade == "SC":
                subjectWiseSelection[sub]["SC"][usn] = studentJson[usn]["Name"]
            elif grade == "F":
                subjectWiseSelection[sub]["F"][usn] = studentJson[usn]["Name"]
            else:
                subjectWiseSelection[sub]["AB"][usn] = studentJson[usn]["Name"]

    # pprint.pprint(subjectWiseSelection, indent=4, sort_dicts=False)
    for sub in sheetNames:
        # access a subject's worksheet
        currentWB = workBook[sub]
        # define the cells of headings
        # headingCells = ["A5", "A6", "B6", "D5", "D6", "E6", "G5", "G6", "H6", "J5", "J6", "K6", "M5", "M6", "N6"]
        headingCells = ["A1", "A2", "B2", "D1", "D2", "E2", "G1", "G2", "H2", "J1", "J2", "K2", "M1", "M2", "N2"]
        # assign headers to correct cells
        currentWB["A1"] = "FCD"
        currentWB["A2"] = "USN"
        currentWB["B2"] = "NAME"
        currentWB["D1"] = "FC"
        currentWB["D2"] = "USN"
        currentWB["E2"] = "NAME"
        currentWB["G1"] = "SC"
        currentWB["G2"] = "USN"
        currentWB["H2"] = "NAME"
        currentWB["J1"] = "F"
        currentWB["J2"] = "USN"
        currentWB["K2"] = "NAME"
        currentWB["M1"] = "AB"
        currentWB["M2"] = "USN"
        currentWB["N2"] = "NAME"
        # format header cells (Alignment, border, bold)
        for cell in headingCells:
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            currentWB[cell].font = openpyxl.styles.Font(bold=True)
        # adjust column width of USN column and align all values to the center
        for col in ["A", "D", "G", "J", "M"]:
            currentWB.column_dimensions[col].width = 18.00
        # adjust column width of Name column and align all values to the center
        for col in ["B", "E", "H", "K", "N"]:
            currentWB.column_dimensions[col].width = 30.00
        # adjust column width for spacing columns
        for col in ["C", "I", "F", "L"]:
            currentWB.column_dimensions[col].width = 5.00
        # merge heading cells
        currentWB.merge_cells("A1:B1")
        currentWB.merge_cells("D1:E1")
        currentWB.merge_cells("G1:H1")
        currentWB.merge_cells("J1:K1")
        currentWB.merge_cells("M1:N1")
        # columns where the USN and name of each grade are listed
        columnDict = {
            "FCD": {"usn": "A", "name": "B", "usnLastRow": 3, "nameLastRow": 3},
            "FC": {"usn": "D", "name": "E", "usnLastRow": 3, "nameLastRow": 3},
            "SC": {"usn": "G", "name": "H", "usnLastRow": 3, "nameLastRow": 3},
            "F": {"usn": "J", "name": "K", "usnLastRow": 3, "nameLastRow": 3},
            "AB": {"usn": "M", "name": "N", "usnLastRow": 3, "nameLastRow": 3}
        }
        highestRow = 0
        for grade in columnDict.keys():
            usnGradeList = list(subjectWiseSelection[sub][grade].keys())
            if len(usnGradeList) != 0:
                for row in range(0, len(usnGradeList)):
                    usnCell = columnDict[grade]["usn"] + str(row + 3)
                    currentWB[usnCell] = usnGradeList[row]
                    columnDict[grade]["usnLastRow"] = row + 3
                    currentWB[usnCell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    currentWB[usnCell].border = Border(left=Side(style='thin'),
                                                       right=Side(style='thin'),
                                                       top=Side(style='thin'),
                                                       bottom=Side(style='thin'))
                    nameCell = columnDict[grade]["name"] + str(row + 3)
                    currentWB[nameCell] = subjectWiseSelection[sub][grade][usnGradeList[row]]
                    columnDict[grade]["nameLastRow"] = row + 3
                    currentWB[nameCell].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                    currentWB[nameCell].border = Border(left=Side(style='thin'),
                                                        right=Side(style='thin'),
                                                        top=Side(style='thin'),
                                                        bottom=Side(style='thin'))
            totalLabelCell = str(columnDict[grade]["usn"]) + str(columnDict[grade]["usnLastRow"] + 2)
            currentWB[totalLabelCell] = "TOTAL"
            currentWB[totalLabelCell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[totalLabelCell].border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
            currentWB[totalLabelCell].font = openpyxl.styles.Font(bold=True)
            columnDict[grade]["usnLastRow"] += 2
            totalCell = str(columnDict[grade]["name"]) + str(columnDict[grade]["nameLastRow"] + 2)
            currentWB[totalCell] = len(usnGradeList)
            currentWB[totalCell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[totalCell].border = Border(left=Side(style='thin'),
                                                 right=Side(style='thin'),
                                                 top=Side(style='thin'),
                                                 bottom=Side(style='thin'))
            currentWB[totalCell].font = openpyxl.styles.Font(bold=True)
            columnDict[grade]["nameLastRow"] += 2
            if columnDict[grade]["nameLastRow"] > highestRow:
                highestRow = columnDict[grade]["nameLastRow"]
            if columnDict[grade]["usnLastRow"] > highestRow:
                highestRow = columnDict[grade]["usnLastRow"]

        highestRow += 2
        formatCellsList = []
        currentWB[str("A" + str(highestRow))] = "TOTAL APPEARED"
        formatCellsList.append(str("A" + str(highestRow)))
        currentWB[str("B" + str(highestRow))] = subjectJson[sub]["TOTAL_APPEARED"]
        formatCellsList.append(str("B" + str(highestRow)))

        currentWB[str("D" + str(highestRow))] = "TOP 1"
        currentWB.merge_cells(str("D" + str(highestRow) + ":D" + str(highestRow + 1)))
        formatCellsList.append(str("D" + str(highestRow)))
        topUSN = subjectJson[sub]["TOP_1"]
        currentWB[str("E" + str(highestRow))] = topUSN
        formatCellsList.append(str("E" + str(highestRow)))
        formatCellsList.append(str("D" + str(highestRow + 1)))
        currentWB[str("E" + str(highestRow + 1))] = studentJson[topUSN]["Name"]
        formatCellsList.append(str("E" + str(highestRow + 1)))

        currentWB[str("G" + str(highestRow))] = "TOP 2"
        currentWB.merge_cells(str("G" + str(highestRow) + ":G" + str(highestRow + 1)))
        formatCellsList.append(str("G" + str(highestRow)))
        topUSN = subjectJson[sub]["TOP_2"]
        currentWB[str("H" + str(highestRow))] = topUSN
        formatCellsList.append(str("H" + str(highestRow)))
        formatCellsList.append(str("G" + str(highestRow + 1)))
        currentWB[str("H" + str(highestRow + 1))] = studentJson[topUSN]["Name"]
        formatCellsList.append(str("H" + str(highestRow + 1)))

        currentWB[str("J" + str(highestRow))] = "TOP 3"
        currentWB.merge_cells(str("J" + str(highestRow) + ":J" + str(highestRow + 1)))
        formatCellsList.append(str("J" + str(highestRow)))
        topUSN = subjectJson[sub]["TOP_3"]
        currentWB[str("K" + str(highestRow))] = topUSN
        formatCellsList.append(str("K" + str(highestRow)))
        formatCellsList.append(str("J" + str(highestRow + 1)))
        currentWB[str("K" + str(highestRow + 1))] = studentJson[topUSN]["Name"]
        formatCellsList.append(str("K" + str(highestRow + 1)))

        for cell in formatCellsList:
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            currentWB[cell].font = openpyxl.styles.Font(bold=True)

        try:
            currentWB = workBook["Consolidated"]
        except:
            currentWB = workBook.create_sheet("Consolidated")
        currentWB["B2"] = "Grade/Subject"
        currentWB["B3"] = "FCD"
        currentWB["B4"] = "FC"
        currentWB["B5"] = "SC"
        currentWB["B6"] = "F"
        currentWB["B7"] = "AB"
        currentWB["B9"] = "TOTAL STUDENTS"
        currentWB.column_dimensions["B"].width = 20.00
        headList = ["B2", "B3", "B4", "B5", "B6", "B7", "B9"]
        for cell in headList:
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            currentWB[cell].font = openpyxl.styles.Font(bold=True)
        subList = list(subjectJson.keys())
        colList = []
        for i in range(0, len(subList)):
            colList.append(chr(66 + (i + 1)))
        # noinspection PyUnusedLocal
        rowDict = {"FCD": 3, "FC": 4, "SC": 5, "F": 6, "AB": 7, "TOTAL": 9}
        for num in range(0, len(subList)):
            currentWB[str(colList[num] + str(2))] = subList[num]
            currentWB[str(colList[num] + str(2))].alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                        vertical='center')
            currentWB[str(colList[num] + str(2))].border = Border(left=Side(style='thin'),
                                                                  right=Side(style='thin'),
                                                                  top=Side(style='thin'),
                                                                  bottom=Side(style='thin'))
            currentWB[str(colList[num] + str(2))].font = openpyxl.styles.Font(bold=True)
        for num in range(0, len(subList)):
            cell = str(colList[num] + str(3))
            currentWB.column_dimensions[colList[num]].width = 20.00
            currentWB[cell] = subjectJson[subList[num]]["FCD"]
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            cell = str(colList[num] + str(4))
            currentWB[cell] = subjectJson[subList[num]]["FC"]
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            cell = str(colList[num] + str(5))
            currentWB[cell] = subjectJson[subList[num]]["SC"]
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            cell = str(colList[num] + str(6))
            currentWB[cell] = subjectJson[subList[num]]["F"]
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            cell = str(colList[num] + str(7))
            currentWB[cell] = subjectJson[subList[num]]["AB"]
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
            cell = str(colList[num] + str(9))
            currentWB[cell] = subjectJson[subList[num]]["TOTAL_APPEARED"]
            currentWB[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWB[cell].border = Border(left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin'))
        # save the workbook with a name
        workBook.save(analysisFilePath)
    return 1


def main():
    global fileDir, fileName, sheetName, analysisFilePath, outputDir
    if setFilePath() == 1:
        print("File:", fileName)
        print("Directory:", fileDir)
        getSheetName()
        print("Using sheet \"", sheetName, "\"")
    if toJson() == 1:
        print("Student Marks JSON File:", path.join(outputDir, "studentMarks.json").replace("\\", "/"))
        print("Subject Analysis JSON File:", path.join(outputDir, "subjectAnalysis.json").replace("\\", "/"))
    if markToXl() == 1:
        print("Analysis Excel WorkBook:", analysisFilePath)
    if listStudents() == 1:
        if ciaReport.main(filePath, sheetName, outputDir) == 1:
            print("Program Execution completed !")


if __name__ == "__main__":
    main()
