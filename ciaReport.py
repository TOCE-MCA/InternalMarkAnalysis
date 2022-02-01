import json
import openpyxl
from openpyxl.styles.borders import Border, Side
import os.path as path
import os
import pandas as pd
# noinspection PyUnresolvedReferences
import pprint
from datetime import datetime
import shutil

fileDir = ""
fileName = ""
filePath = ""
sheetName = ""
reportFilePath = ""
outputDirectory = ""
xlFile = ""


# noinspection PyArgumentList
def setFilePath(passedFilePath, localSheetName):
    global fileDir, fileName, filePath, reportFilePath, sheetName, outputDirectory
    filePath = passedFilePath.replace("\\", "/")
    fileName = path.basename(filePath)
    fileDir = path.dirname(filePath)
    sheetName = localSheetName
    reportFileName = "CIA_" + str(sheetName).replace(" ", "") + "_Reports.xlsx"
    reportFilePath = path.join(outputDirectory, reportFileName).replace("\\", "/")
    return 1


def checkSheetName():
    global filePath, fileName, fileDir
    wb = openpyxl.open(filePath, read_only=True, keep_links=False)
    sheetNames = list(wb.sheetnames)
    if "ExamDetails" not in sheetNames:
        print("\"ExamDetails\" sheet not found !\n CIA report could not be generated !\n")
        exit(-1)
    else:
        print("\"ExamDetails\" sheet exists")


def checkTemplate():
    if not path.exists(path.join(os.curdir, "CIA_template.xlsx").replace("\\", "/")):
        print("Report template file missing in source directory")
        exit(-1)
    else:
        print("Report template exists")


# noinspection PyBroadException,PyTypeChecker
def toJson():
    global fileDir, filePath
    try:
        internalFrame = pd.read_excel(filePath, sheet_name="ExamDetails", header=0, usecols="A:G")
    except:
        internalFrame = pd.read_excel(filePath, sheet_name="ExamDetails", header=0, usecols="A:G", engine="openpyxl")

    examDetails = {}
    for sub in internalFrame["SubCode"]:
        if sub not in examDetails.keys():
            examDetails[sub] = {"section": {}}

    for i in range(0, len(internalFrame["SubCode"])):
        examDetails[internalFrame["SubCode"][i]]["name"] = internalFrame["SubjectName"][i]
        examDetails[internalFrame["SubCode"][i]]["year"] = internalFrame["Year"][i]
        examDetails[internalFrame["SubCode"][i]]["semester"] = internalFrame["Semester"][i]
        tempExamDate = datetime.strptime(str(internalFrame["DateOfExam"][i]), "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
        examDetails[internalFrame["SubCode"][i]]["dateOfExam"] = str(tempExamDate)
        examDetails[internalFrame["SubCode"][i]]["section"][internalFrame["Section"][i]] = internalFrame["FacultyName"][
            i]

    with open(path.join(outputDirectory, "examDetails.json").replace("\\", "/"), "w") as file:
        json.dump(examDetails, file, indent=4, sort_keys=False)


def getAboveClassAverage(marksJson, classAverage, sheet):
    if "_" in sheet:
        sub = sheet.split("_")[0]
        section = sheet.split("_")[1].strip()
    else:
        sub = sheet
        section = "E"
    count = 0
    for usn in marksJson.keys():
        if sub in marksJson[usn]["marks"].keys():
            if section == "E":
                if marksJson[usn]["marks"][sub]["marks"] >= classAverage:
                    count += 1
            else:
                if marksJson[usn]["Section"] == section:
                    if marksJson[usn]["marks"][sub]["marks"] >= classAverage:
                        count += 1
    return count


# noinspection PyShadowingBuiltins
def generateExcelReport():
    global reportFilePath, fileDir, sheetName
    shutil.copy("CIA_template.xlsx", reportFilePath)

    reportWB = openpyxl.load_workbook(reportFilePath, read_only=False)
    templateSheet = reportWB["template"]

    ia = str(sheetName).split("-")[1].strip()
    if ia == str(1):
        templateSheet["B5"] = "CIA - I - Marks Statement"
    elif ia == str(2):
        templateSheet["B5"] = "CIA - II - Marks Statement"
    elif ia == str(3):
        templateSheet["B5"] = "CIA - III - Marks Statement"

    with open(path.join(outputDirectory, "examDetails.json").replace("\\", "/")) as file:
        examDetails = json.load(file)

    for sub in examDetails.keys():
        if "E" in list(examDetails[sub]["section"].keys()):
            targetSheet = reportWB.copy_worksheet(templateSheet)
            targetSheet.title = str(sub).strip(" ")
        elif "A" in list(examDetails[sub]["section"].keys()):
            targetSheet = reportWB.copy_worksheet(templateSheet)
            targetSheet.title = str(str(sub).strip(" ") + "_A")
        if "B" in list(examDetails[sub]["section"].keys()):
            targetSheet = reportWB.copy_worksheet(templateSheet)
            targetSheet.title = str(str(sub).strip(" ") + "_B")
    reportWB.remove(templateSheet)

    with open(path.join(outputDirectory, "studentMarks.json").replace("\\", "/")) as file:
        marksJson = json.load(file)
    indexDict = {}
    subjectMarks = {}
    for sheet in reportWB.sheetnames:
        indexDict[sheet] = {"serial": 0, "start": 13, "end": 13, "absent": 0, "fail": 0, "sum": 0, "max": 0, "min": 0}
        marksList = []
    for usn in marksJson:
        # noinspection PyArgumentList
        for sub in marksJson[usn]["marks"].keys():
            try:
                if "E" not in list(examDetails[sub]["section"].keys()):
                    if marksJson[usn]["Section"] == "A":
                        tempSheetName = str(sub).strip(" ") + "_A"
                    else:
                        tempSheetName = str(sub).strip(" ") + "_B"
                else:
                    tempSheetName = str(sub).strip(" ")
                if tempSheetName not in subjectMarks.keys():
                    subjectMarks[tempSheetName] = []
            except KeyError:
                print("Check Subject code in current working sheet!, Enter only subject codes.\n"
                      "Subject code in this sheet should match the subject codes in \"ExamDetails\" sheet.")

            indexDict[tempSheetName]["serial"] += 1
            indexDict[tempSheetName]["end"] += 1
            currentWorkSheet = reportWB[tempSheetName]
            currentWorkSheet.insert_rows(idx=int(indexDict[tempSheetName]["end"]), amount=1)
            cell = "A" + str(indexDict[tempSheetName]["end"])
            currentWorkSheet[cell] = indexDict[tempSheetName]["serial"]
            currentWorkSheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWorkSheet[cell].border = Border(left=Side(style='thin'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))
            cell = "B" + str(indexDict[tempSheetName]["end"])
            currentWorkSheet[cell] = usn
            currentWorkSheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWorkSheet[cell].border = Border(left=Side(style='thin'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))
            cell = "C" + str(indexDict[tempSheetName]["end"])
            currentWorkSheet[cell] = marksJson[usn]["Name"]
            currentWorkSheet[cell].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            currentWorkSheet[cell].border = Border(left=Side(style='thin'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))
            cell = "D" + str(indexDict[tempSheetName]["end"])
            if marksJson[usn]["marks"][sub]["marks"] != -1:
                currentWorkSheet[cell] = marksJson[usn]["marks"][sub]["marks"]
                indexDict[tempSheetName]["sum"] += marksJson[usn]["marks"][sub]["marks"]
                # noinspection PyUnboundLocalVariable
                marksList.append(marksJson[usn]["marks"][sub]["marks"])
                subjectMarks[tempSheetName].append(marksJson[usn]["marks"][sub]["marks"])
                indexDict[tempSheetName]["max"] = max(subjectMarks[tempSheetName])
                indexDict[tempSheetName]["min"] = min(subjectMarks[tempSheetName])
            else:
                currentWorkSheet[cell] = "AB"
                indexDict[tempSheetName]["absent"] += 1
            currentWorkSheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWorkSheet[cell].border = Border(left=Side(style='thin'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))
            cell = "E" + str(indexDict[tempSheetName]["end"])
            if marksJson[usn]["marks"][sub]["percentage"] != "N/A":
                currentWorkSheet[cell] = marksJson[usn]["marks"][sub]["percentage"]
                currentWorkSheet[cell].number_format = "0.00"
            else:
                currentWorkSheet[cell] = "AB"
            currentWorkSheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWorkSheet[cell].border = Border(left=Side(style='thin'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))
            if marksJson[usn]["marks"][sub]["grade"] == "F":
                indexDict[tempSheetName]["fail"] += 1

    # header
    for sheet in reportWB.sheetnames:
        currentWorkSheet = reportWB[sheet]
        if "_" in sheet:
            sub = sheet.split("_")[0]
            section = sheet.split("_")[1]
        else:
            sub = sheet
            section = "E"
        # subject code
        currentWorkSheet["B7"] = sub
        # subject name
        currentWorkSheet["B8"] = examDetails[sub]["name"]
        # year / sem
        cellValue = str(examDetails[sub]["year"] + "/ " + examDetails[sub]["semester"])
        currentWorkSheet["B9"] = cellValue
        # Date of exam
        currentWorkSheet["B10"] = examDetails[sub]["dateOfExam"]
        # faculty
        currentWorkSheet["B11"] = examDetails[sub]["section"][section]

    # calculations
    for sheet in reportWB.sheetnames:
        currentWorkSheet = reportWB[sheet]
        # total students
        cell = "E" + str(indexDict[sheet]["end"] + 2)
        currentWorkSheet[cell] = indexDict[sheet]["serial"]
        # present
        cell = "E" + str(indexDict[sheet]["end"] + 3)
        currentWorkSheet[cell] = indexDict[sheet]["serial"] - indexDict[sheet]["absent"]
        # pass
        cell = "E" + str(indexDict[sheet]["end"] + 4)
        currentWorkSheet[cell] = indexDict[sheet]["serial"] - indexDict[sheet]["fail"] - indexDict[sheet]["absent"]
        # fail
        cell = "E" + str(indexDict[sheet]["end"] + 5)
        currentWorkSheet[cell] = indexDict[sheet]["fail"]
        # class average
        cell = "E" + str(indexDict[sheet]["end"] + 6)
        classAverage = round(indexDict[sheet]["sum"] / (indexDict[sheet]["serial"] - indexDict[sheet]["absent"]), 2)
        currentWorkSheet[cell] = classAverage
        # above class average
        cell = "E" + str(indexDict[sheet]["end"] + 7)
        count = getAboveClassAverage(marksJson, classAverage, sheet)
        currentWorkSheet[cell] = count
        # max marks
        cell = "E" + str(indexDict[sheet]["end"] + 8)
        currentWorkSheet[cell] = indexDict[sheet]["max"]
        # min marks
        cell = "E" + str(indexDict[sheet]["end"] + 9)
        currentWorkSheet[cell] = indexDict[sheet]["min"]
        # pass percentage
        cell = "E" + str(indexDict[sheet]["end"] + 10)
        percent = ((indexDict[sheet]["serial"] - indexDict[sheet]["fail"]) / (indexDict[sheet]["serial"]))*100
        percent = round(percent, 2)
        currentWorkSheet[cell] = percent
        # teaching effectiveness [(class avg/total marks)*100]
        cell = "C" + str(indexDict[sheet]["end"] + 11)
        cellFormula = "=(E" + str(indexDict[sheet]["end"] + 6) + "/40)*100"
        currentWorkSheet[cell] = cellFormula
        currentWorkSheet[cell].font = openpyxl.styles.Font(name="Times New Roman", bold=True, size=11)
        # learning effectiveness [(no. of students above class average/total students)*100]
        cell = "C" + str(indexDict[sheet]["end"] + 12)
        cellFormula = "=(E" + str(indexDict[sheet]["end"] + 7) + "/E" + str(indexDict[sheet]["end"] + 4) + ")*100"
        currentWorkSheet[cell] = cellFormula
        currentWorkSheet[cell].font = openpyxl.styles.Font(name="Times New Roman", bold=True, size=11)

    # footer
    for sheet in reportWB.sheetnames:
        currentWorkSheet = reportWB[sheet]
        for i in range(2, 11):
            cell_1 = "A" + str(indexDict[sheet]["end"] + i)
            cell_2 = "D" + str(indexDict[sheet]["end"] + i)
            cellRange = cell_1 + ":" + cell_2
            currentWorkSheet.merge_cells(cellRange)
        for i in range(11, 13):
            cell_1 = "C" + str(indexDict[sheet]["end"] + i)
            cell_2 = "E" + str(indexDict[sheet]["end"] + i)
            cellRange = cell_1 + ":" + cell_2
            currentWorkSheet.merge_cells(cellRange)
            currentWorkSheet[cell_1].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            currentWorkSheet.row_dimensions[indexDict[sheet]["end"] + i].height = 25.0

        principalFooterStart = "A" + str(indexDict[sheet]["end"] + 17)
        principalFooterStop = "B" + str(indexDict[sheet]["end"] + 17)
        cellRange = principalFooterStart + ":" + principalFooterStop
        currentWorkSheet.merge_cells(cellRange)
        currentWorkSheet[principalFooterStart].alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                     vertical='center')
        currentWorkSheet[principalFooterStart].font = openpyxl.styles.Font(name="Times New Roman", bold=True, size=11)

        cell = "C" + str(indexDict[sheet]["end"] + 17)
        currentWorkSheet[cell].alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                     vertical='center')
        currentWorkSheet[cell].font = openpyxl.styles.Font(name="Times New Roman", bold=True, size=11)

        staffFooterStart = "D" + str(indexDict[sheet]["end"] + 17)
        staffFooterStop = "E" + str(indexDict[sheet]["end"] + 17)
        cellRange = staffFooterStart + ":" + staffFooterStop
        currentWorkSheet.merge_cells(cellRange)
        currentWorkSheet[staffFooterStart].alignment = openpyxl.styles.Alignment(horizontal='center',
                                                                                 vertical='center')
        currentWorkSheet[staffFooterStart].font = openpyxl.styles.Font(name="Times New Roman", bold=True, size=11)

    reportWB.save(reportFilePath)


# noinspection PyShadowingNames
def main(passedFilePath, sheetName, outputDir):
    global fileName, fileDir, filePath, reportFilePath
    global outputDirectory
    outputDirectory = outputDir
    setFilePath(passedFilePath, sheetName)
    checkSheetName()
    checkTemplate()
    toJson()
    generateExcelReport()
    print("Reports generated at: ", reportFilePath.replace("\\", "/"))
    return 1
